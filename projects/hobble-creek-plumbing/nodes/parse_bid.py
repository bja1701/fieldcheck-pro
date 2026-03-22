"""
Node 1: parse_bid
Reads the bid sheet (CSV or XLSX/XLSM) and extracts line items.

Output state keys updated:
  bid_items: list of {room, name, bid_qty, unit_price}
"""
from __future__ import annotations

import csv
import json
import os
import re
from collections import Counter
from pathlib import Path

from state import FieldCheckState

# Rows whose item name matches these patterns are metadata, not fixtures
_SKIP_PATTERNS = re.compile(
    r"^(total|subtotal|sub-total|grand total|material|labor|tax|overhead|profit|"
    r"description only|allowance|note|phase)",
    re.IGNORECASE,
)


def _is_skip_row(name: str) -> bool:
    return bool(_SKIP_PATTERNS.match(name.strip()))


def _read_csv(path: Path) -> tuple[list[str], list[list[str]]]:
    with open(path, newline="", encoding="utf-8-sig") as f:
        reader = csv.reader(f)
        rows = [r for r in reader if any(c.strip() for c in r)]
    if not rows:
        return [], []
    return rows[0], rows[1:]


def _parse_item_list(text: str) -> list[tuple[int, str]]:
    """Parse Hobble Creek item text: '2 Water Heater, 1 Floor Drain' → [(2, 'Water Heater'), (1, 'Floor Drain')]."""
    # Split on comma only when followed by a digit (start of next item)
    parts = re.split(r",\s*(?=\d)", text)
    items = []
    for part in parts:
        part = " ".join(part.strip().split())  # normalise whitespace
        m = re.match(r"^(\d+)\s+(.+)$", part)
        if m:
            items.append((int(m.group(1)), m.group(2).strip()))
    return items


def _read_final_estimate(ws) -> tuple[list[str], list[list[str]]]:
    """Read Hobble Creek 'Final Estimate' sheet: alternating room-name / item-text rows.

    Detects the ROOM+ITEM header row, then parses pairs:
      row N  : col_room = room name
      row N+1: col_item = '2 Toilet, 1 Lav, …'
    Returns synthetic headers ['room','name','qty','price'] + data rows so the
    rest of parse_bid can treat it like any other spreadsheet.
    """
    room_col = item_col = header_row = None
    for row in ws.iter_rows():
        cells = {c.column: str(c.value or "").strip() for c in row if c.value}
        uppers = {col: v.upper() for col, v in cells.items()}
        if "ROOM" in uppers.values() and "ITEM" in uppers.values():
            room_col = next(c for c, v in uppers.items() if v == "ROOM")
            item_col = next(c for c, v in uppers.items() if v == "ITEM")
            header_row = row[0].row
            break

    if not header_row:
        return [], []

    data_rows: list[list[str]] = []
    current_room = ""
    for row in ws.iter_rows(min_row=header_row + 1):
        r_val = str(ws.cell(row=row[0].row, column=room_col).value or "").strip()
        i_val = str(ws.cell(row=row[0].row, column=item_col).value or "").strip()

        # Stop at footer sentinels
        if r_val.upper() in ("FIXTURES", "END"):
            break

        if r_val and not i_val:
            current_room = r_val
        elif i_val and current_room:
            for qty, name in _parse_item_list(i_val):
                data_rows.append([current_room, name, str(qty), ""])

    return ["room", "name", "qty", "price"], data_rows


def _read_xlsx(path: Path) -> tuple[list[str], list[list[str]]]:
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)

    # Try every sheet for the Hobble Creek 'Final Estimate' alternating format first
    for sheet_name in wb.sheetnames:
        headers, data_rows = _read_final_estimate(wb[sheet_name])
        if data_rows:
            print(f"  [parse_bid] Detected Final Estimate format in sheet '{sheet_name}'")
            return headers, data_rows

    # Fall back: active sheet, row-per-item format (e.g. Baker CSV-style xlsx)
    ws = wb.active
    rows = [[str(c.value or "") for c in row] for row in ws.iter_rows()]
    rows = [r for r in rows if any(v.strip() for v in r)]
    if not rows:
        return [], []
    return rows[0], rows[1:]


def _heuristic_col_map(headers: list[str]) -> dict | None:
    """Map column headers to (room, name, qty, price) by keyword matching."""
    h = [h.lower().strip() for h in headers]
    result: dict[str, int | None] = {"room": None, "name": None, "qty": None, "price": None}
    for i, col in enumerate(h):
        if col in ("room", "location", "area", "space"):
            result["room"] = i
        elif col in ("item", "description", "fixture", "work", "name"):
            result["name"] = i
        elif col in ("qty", "quantity", "count", "units", "ea", "ea.", "#"):
            result["qty"] = i
        elif col in ("price", "unit price", "unit_price", "cost", "rate", "amount"):
            result["price"] = i
    if result["room"] is not None and result["name"] is not None and result["qty"] is not None:
        return result
    return None


def _col_map_with_llm(headers: list[str], sample: list[list[str]]) -> dict:
    """Fall back to Gemini to identify columns when heuristics fail."""
    api_key = os.environ.get("GOOGLE_API_KEY") or os.environ.get("GEMINI_API_KEY", "")
    if not api_key:
        raise RuntimeError("GOOGLE_API_KEY not set — cannot identify bid columns")
    from google import genai
    client = genai.Client(api_key=api_key)
    sample_text = "\n".join("  " + " | ".join(r) for r in sample[:5])
    prompt = (
        f"Contractor bid sheet column headers: {headers}\n"
        f"Sample rows:\n{sample_text}\n\n"
        f"Return ONLY JSON identifying zero-based column index for room, name, qty, price "
        f"(use null if not present): {{\"room\":0,\"name\":1,\"qty\":2,\"price\":3}}"
    )
    model = os.environ.get("GEMINI_MODEL", "gemini-2.5-flash")
    resp = client.models.generate_content(model=model, contents=prompt)
    text = resp.text.strip()
    start, end = text.find("{"), text.rfind("}") + 1
    return json.loads(text[start:end])


def _deduplicate_room_names(items: list[dict]) -> list[dict]:
    """Number rooms that appear in multiple non-contiguous groups.

    E.g. bid sheet with two separate '1/2 Bath' sections becomes
    '1/2 Bath 1' and '1/2 Bath 2'. Single-occurrence rooms unchanged.
    """
    if not items:
        return items

    # Build contiguous segments: each time room name changes, start a new segment
    segments: list[tuple[str, int, int]] = []  # (room, start_idx, end_idx exclusive)
    seg_room = items[0]["room"]
    seg_start = 0
    for i, item in enumerate(items[1:], 1):
        if item["room"] != seg_room:
            segments.append((seg_room, seg_start, i))
            seg_room, seg_start = item["room"], i
    segments.append((seg_room, seg_start, len(items)))

    # Which room names appear more than once as separate segments?
    multi = {name for name, cnt in Counter(s[0] for s in segments).items() if cnt > 1}
    if not multi:
        return items

    seg_counter: dict[str, int] = {}
    result = [dict(it) for it in items]
    for room_name, start, end in segments:
        if room_name in multi:
            seg_counter[room_name] = seg_counter.get(room_name, 0) + 1
            new_name = f"{room_name} {seg_counter[room_name]}"
            for i in range(start, end):
                result[i]["room"] = new_name
    return result


def parse_bid(state: FieldCheckState) -> dict:
    path = Path(state["bid_path"])
    ext = path.suffix.lower()

    if ext == ".csv":
        headers, data_rows = _read_csv(path)
    elif ext in (".xlsx", ".xlsm", ".xls"):
        headers, data_rows = _read_xlsx(path)
    else:
        return {"error": f"Unsupported bid file type: {ext}", "bid_items": []}

    col_map = _heuristic_col_map(headers)
    if col_map is None:
        col_map = _col_map_with_llm(headers, data_rows)

    room_col = col_map.get("room")
    name_col = col_map.get("name")
    qty_col = col_map.get("qty")
    price_col = col_map.get("price")

    if room_col is None or name_col is None or qty_col is None:
        return {"error": "Could not identify room/name/qty columns in bid sheet", "bid_items": []}

    items = []
    current_room = ""
    for row in data_rows:
        if not row or not any(r.strip() for r in row):
            continue
        name = row[name_col].strip() if name_col < len(row) else ""
        if not name or _is_skip_row(name):
            continue

        # Room column may be blank — carry forward the last non-blank room
        room_val = row[room_col].strip() if room_col < len(row) else ""
        if room_val:
            current_room = room_val

        try:
            qty = float(row[qty_col].strip().replace(",", "") if qty_col < len(row) else "0") if row[qty_col].strip() else 0
        except ValueError:
            qty = 0

        price = None
        if price_col is not None and price_col < len(row):
            raw_price = row[price_col].strip().replace("$", "").replace(",", "")
            try:
                price = float(raw_price)
            except ValueError:
                pass

        items.append({
            "room": current_room,
            "name": name,
            "bid_qty": int(qty),
            "unit_price": price,
        })

    items = _deduplicate_room_names(items)

    return {
        "bid_items": items,
        "logs": [f"parse_bid: {len(items)} items from {path.name}"],
    }
